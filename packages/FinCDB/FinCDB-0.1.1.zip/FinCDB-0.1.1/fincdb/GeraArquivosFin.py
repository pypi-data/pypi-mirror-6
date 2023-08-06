# -*- coding: UTF-8 -*-

#-------------------------------------------------------------------------------
# Name:        GeraArquivosFin.py
# Purpose:     Gerar os arquivos necessarios para o projeto Finanpy
#               (datas, cdi, etc)
#
# Author:      MGFac
#
# Created:     24/08/2013
# Copyright:   (c) MGFac 2013
# Licence:     <your licence>
#-------------------------------------------------------------------------------
import os
import FinDt
import pickle
import openpyxl as xl
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter


def CarregaCDI(file, range):
    try:
        wb = load_workbook(file)
        ws = wb.get_active_sheet()
        dicCDI = { FinDt.DatasFinanceiras.DateToStr(FinDt.DatasFinanceiras, row[0].value):row[4].value for row in ws.range(range)}
        return dicCDI
    except IOError as Ierr:
        print("Erro de Leitura: ", str(Ierr))

def PickleDic(filename, dicionario):
    try:
        with open(filename, 'wb') as wfile:
            pickle.dump(dicionario, wfile)
    except IOError as Ierr:
        print("Erro de Leitura: ", str(Ierr))

def UnPickleDic(picklefile):
    try:
        with open(picklefile, 'rb') as rfile:
            return pickle.load(rfile)
    except IOError as Ierr:
        print("Erro de Leitura: ", str(Ierr))


def main():
    pkCDI = 'CDICetip.pkl'
    xlCDI = 'DiCetip.xlsm'
    xlRange = 'A1:E6861'

    os.chdir('C:\\Apostilas\\Python\\BCB\\ProjFinanPy')
    PickleDic(pkCDI, CarregaCDI(xlCDI, xlRange))


if __name__ == '__main__':
    main()



