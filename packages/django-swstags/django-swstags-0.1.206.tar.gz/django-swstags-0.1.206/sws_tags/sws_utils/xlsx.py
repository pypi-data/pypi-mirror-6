#!/usr/bin/env python
# encoding: utf-8

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from sws_tags.sws_utils import json_encode

import csv
from django.http import HttpResponse
from django.template import loader, Context
from decimal import Decimal
from datetime import date, datetime, timedelta
import time
import ujson
# import datetime



# FOR EXCEL GENERATING
# import xlwt
# ezxf = xlwt.easyxf

from sws_tags.sws_utils.common_utils import *
from sws_tags.sws_utils.messages import *
from sws_tags.sws_utils.cube import *

import traceback



# FOR EXCEL 
try:
	import cStringIO as StringIO
except ImportError:
	import StringIO

from django.http import HttpResponse

from xlsxwriter.workbook import Workbook




# FOR EXCEL GENERATING
import xlwt
ezxf = xlwt.easyxf

from sws_tags.sws_utils.common_utils import *
from sws_tags.sws_utils.messages import *
from sws_tags.sws_utils.cube import *

import traceback


# import xlwt workbook = xlwt.Workbook(encoding = 'ascii') 
# worksheet = workbook.add_sheet('My Worksheet') 
# worksheet.write(0, 0, label = 'Row 0, Column 0 Value') 
# workbook.save('Excel_Workbook.xls') 

def ExportFile(param_export_file):
	def read_and_flush():
		output.seek(0)
		data = output.read()
		output.seek(0)
		output.truncate()
		return data


	def writeInStream():
		row = 0
		col = 0
		for h in headers:
			sheet.write(row, col, h,bold)
			col += 1
		row = 1
		col = 0
		for trr in queryset:
			for c in col_name:
				if type(trr[c]) == datetime:
					try:
						trr[c]=param_export_file['request_data']['django_timezone'].normalize(trr[c]).strftime('%Y-%m-%d %H:%M:%S')
					except:
						# print traceback.format_exc()
						trr[c]=str(trr[c])
				sheet.write(row, col, trr[c])
				col += 1
			row += 1
			col = 0
			data = read_and_flush()
			yield data

		book.close()

		data = read_and_flush()
		yield data	
	if param_export_file['format'] == 'excel':
		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'

		queryset = param_export_file['queryset']
		headers = param_export_file['headers']
		col_name = param_export_file['col_name']
		filename = param_export_file['filename']
		logger = param_export_file['logger']

		output = StringIO.StringIO()

		book = Workbook(output)
		sheet = book.add_worksheet('test') 

		bold = book.add_format({'bold': 1})

		response = HttpResponse(writeInStream(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
		response['Content-Disposition'] = "attachment; filename="+ filename
		return response

	elif param_export_file['format'] == 'csv':
		param_export_file['filename'] = param_export_file['filename']+'_'+str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))

		json_data = get_csv_query(param_export_file['request_data'], param_export_file['queryset'], param_export_file['filename'])
		response = HttpResponse(json_data, mimetype='text/csv') 
		param_export_file['filename'] += '.csv'
		response['Content-Disposition'] = 'attachment; filename=' + param_export_file['filename']
		return response


def get_csv_query(request, queryset, filename, col_order = None):

	kind_to_xf_map = {
		'date': ezxf(num_format_str='yyyy-mm-dd HH:MM'),
		'int': ezxf(num_format_str='#,##0'),
		'money': ezxf('font: italic on; pattern: pattern solid, fore-colour grey25',
			num_format_str='€#,##0.00'),
		'price': ezxf(num_format_str='#0.000'),
		'text': ezxf(),
		'boolean': ezxf(),
	}

	try:
		# csv_file = open(filename,"w")
		csv_file = HttpResponse()

		csv_writer = csv.writer(csv_file, dialect='excel', quoting=csv.QUOTE_MINIMAL, delimiter='|')

		if type(queryset)!=list:
			q = queryset[0]

			
			fields = []
			for f in queryset._fields:
				fields.append(q[f])
			for f in queryset.aggregate_names:
				fields.append(q[f])

			fields_name = []
			for f in queryset._fields:
				fields_name.append(unicode(getFieldVerboseName(queryset,f)))
			for f in queryset.aggregate_names:
				fields_name.append(f)


			data_xfs_types_used = [k for k in get_field_types(fields)]
			data_xfs = [kind_to_xf_map[k] for k in get_field_types(fields)]

			data = []


			for i, q in enumerate(queryset):
				aux = []
				i=0
				for f in queryset._fields:

					if data_xfs_types_used[i] == 'date':
						try:
							date_normalize=request['django_timezone'].normalize(q[f]).strftime('%Y-%m-%d %H:%M:%S')
							aux.append(str(date_normalize))
						except:
							# print traceback.format_exc()
							aux.append(q[f])

					else:
						aux.append(q[f])
					i+=1
				for f in queryset.aggregate_names:
					aux.append(q[f])
				data.append(aux)

		else:
			fields_name = []
			if col_order:
				fields_name=col_order
			else:
				for k,v in queryset[0].items():
					fields_name.append(unicode(k))

			data = []

			for q in queryset:
				v_data=[]
				for k in fields_name:
					# print 'excel-->',q[k],'--',k
					v_data.append(q[k])
				data.append(v_data)

		csv_writer.writerow(fields_name)

		for row in data:
			csv_writer.writerow(row)							
								
		# csv_file.close()
		return csv_file

	except Exception, e:
		swslog('error','get_csv_query',e)
		return False


def get_field_types(fields):
	# fields = self.fields
	field_types = []
	prev_len = 0
	found = False
	for lf in fields:
		found = False
		if type(lf) == unicode:
			field_types.append('text')
			found = True
		elif type(lf) == long:
			field_types.append('int')
			found = True
		elif type(lf) == int:
			field_types.append('int')
			found = True
		elif type(lf) == bool:
			field_types.append('boolean')
			found = True
		elif type(lf) == float: 
			field_types.append('price')
			found = True
		elif type(lf) == Decimal: 
			field_types.append('price')
			found = True                    
		elif type(lf) == date:
			field_types.append('date')
			found = True                
		elif type(lf) == datetime: 
			field_types.append('date')
			found = True              
		if found is not True: 
			field_types.append('text')
	return field_types












# def ExportFile_STYLE(param_export_file): # STYLE

# 	if param_export_file['format'] == 'excel':

# 		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'

# 		queryset = param_export_file['queryset']
# 		headers = param_export_file['headers']
# 		col_name = param_export_file['col_name']
# 		filename = param_export_file['filename']
# 		logger = param_export_file['logger']


# 		# styles = dict(
# 		#     bold = 'font: bold 1',
# 		#     italic = 'font: italic 1',
# 		#     # Wrap text in the cell
# 		#     wrap_bold = 'font: bold 1; align: wrap 1;',
# 		#     # White text on a blue background
# 		#     reversed = 'pattern: pattern solid, fore_color blue; font: color white;',
# 		#     # Light orange checkered background
# 		#     light_orange_bg = 'pattern: pattern fine_dots, fore_color white, back_color orange; font: bold 1; align: wrap 1;',
# 		#     # Heavy borders
# 		#     bordered = 'border: top thick, right thick, bottom thick, left thick;',
# 		#     # 16 pt red text
# 		#     big_red = 'font: height 320, color orange;',
# 		# )



# 		book = xlwt.Workbook()
# 		sheet = book.add_sheet('file_xlsx')
# 		style = xlwt.easyxf('pattern: pattern fine_dots, fore_color white, back_color orange; font: bold 1; align: wrap 1;')
# 		style_content = xlwt.easyxf('font: bold 1')


# 		# wb = Workbook()
# 		# ws = wb.get_active_sheet()
# 		# ws.title = filename


# 		row = 0
# 		col = 0
# 		for h in headers:
# 			sheet.write(row, col, h, style)
# 			# cell = ws.cell(row = row, column = col)
# 			# cell.value = h
# 			# cell.style = style
# 			# printº 'zzzzzzzzzzzzzzzzz',cell
# 			col += 1
	
# 		row = 1
# 		col = 0
# 		for trr in queryset:
# 			for c in col_name:
# 				# cell = ws.cell(row = row, column = col)
# 				sheet.write(row, col, trr[c])
# 				col += 1
# 				# cell.value = trr[c]

				

# 			row += 1
# 			col = 0

# 		# response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') 		
# 		# response['Content-Disposition'] = 'attachment; filename='+filename
# 		# wb.save(response)

# 		# response = HttpResponse(mimetype="application/ms-excel")
# 		# response['Content-Disposition'] = 'attachment; filename= test.xls'
# 		book.save(response)

# 		return response

# 	elif param_export_file['format'] == 'csv':
# 		param_export_file['filename'] = param_export_file['filename']+'_'+str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))

# 		json_data = get_csv_query(param_export_file['request_data'], param_export_file['queryset'], param_export_file['filename'])
# 		response = HttpResponse(json_data, mimetype='text/csv') 
# 		param_export_file['filename'] += '.csv'
# 		response['Content-Disposition'] = 'attachment; filename=' + param_export_file['filename']
# 		return response


# def ExportFileSaveinstream(param_export_file):
# 	if param_export_file['format'] == 'excel':
# 		param_export_file['filename'] = param_export_file['filename']+'_' + str(datetime.now().strftime("%Y-%m-%d_%H-%M-%S"))+'.xlsx'

# 		queryset = param_export_file['queryset']
# 		headers = param_export_file['headers']
# 		col_name = param_export_file['col_name']
# 		filename = param_export_file['filename']
# 		logger = param_export_file['logger']

# 		response = HttpResponse(mimetype='application/ms-excel') 		
# 		response['Content-Disposition'] = 'attachment; filename='+filename


# 		wb = Workbook()
# 		ws = wb.get_active_sheet()
# 		ws.title = filename

# 		row = 0
# 		col = 0
# 		for h in headers:
# 			cell = ws.cell(row = row, column = col)
# 			cell.value = h
# 			col += 1
	
# 		row = 1
# 		col = 0

# 		wb.save(response)
# 		for trr in queryset:
# 			for c in col_name:
# 				cell = ws.cell(row = row, column = col)
# 				col += 1
# 				cell.value = trr[c]

# 			wb.save(response)
# 			row += 1
# 			col = 0

# 		wb.save(response)
# 		return response